"""
Module d'export Excel pour CamerJob Watch
Génère des fichiers Excel formatés et professionnels
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os
from datetime import datetime


# Couleurs Cameroun
VERT    = "1B5E20"
VERT_C  = "2E7D32"
VERT_L  = "C8E6C9"
ROUGE   = "C62828"
JAUNE   = "F9A825"
GRIS    = "F5F5F5"
GRIS_F  = "EEEEEE"
BLANC   = "FFFFFF"
NOIR    = "212121"
BLEU    = "1565C0"


def style_header(ws, row, cols, bg=VERT, fg=BLANC, taille=11, bold=True):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(name='Calibri', bold=bold, size=taille, color=fg)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        bottom=Side(style='medium', color='000000'),
        top=Side(style='thin', color='888888'),
        left=Side(style='thin', color='888888'),
        right=Side(style='thin', color='888888'),
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def style_row_data(ws, row, cols, alt=False):
    bg = GRIS_F if alt else BLANC
    fill = PatternFill("solid", fgColor=bg)
    font = Font(name='Calibri', size=10, color=NOIR)
    align = Alignment(vertical='top', wrap_text=True)
    border = Border(
        bottom=Side(style='thin', color='DDDDDD'),
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def exporter_excel(offres, type_rapport='toutes'):
    """
    Génère un fichier Excel complet avec :
    - Onglet résumé avec statistiques
    - Onglet offres d'emploi
    - Onglet appels d'offres
    - Graphiques automatiques
    """
    wb = openpyxl.Workbook()

    # ─── ONGLET 1 : TABLEAU DE BORD ───
    ws_dash = wb.active
    ws_dash.title = "📊 Tableau de Bord"
    _creer_onglet_dashboard(ws_dash, offres)

    # ─── ONGLET 2 : OFFRES D'EMPLOI ───
    emplois = [o for o in offres if o.type_offre == 'emploi']
    if emplois:
        ws_emploi = wb.create_sheet("💼 Offres d'Emploi")
        _creer_onglet_emplois(ws_emploi, emplois)

    # ─── ONGLET 3 : APPELS D'OFFRES ───
    appels = [o for o in offres if o.type_offre == 'appel_offre']
    if appels:
        ws_ao = wb.create_sheet("📋 Appels d'Offres")
        _creer_onglet_appels_offres(ws_ao, appels)

    # ─── ONGLET 4 : TOUTES LES OFFRES ───
    ws_all = wb.create_sheet("📋 Toutes les Offres")
    _creer_onglet_toutes(ws_all, offres)

    # Sauvegarder
    exports_dir = os.path.join(os.path.dirname(__file__), '..', 'exports')
    os.makedirs(exports_dir, exist_ok=True)
    nom_fichier = f"CamerJobWatch_{type_rapport}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(exports_dir, nom_fichier)
    wb.save(filepath)
    return filepath


def _creer_onglet_dashboard(ws, offres):
    """Onglet résumé avec statistiques clés"""
    # Largeurs colonnes
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20

    # ── Titre principal ──
    ws.merge_cells('A1:D1')
    ws['A1'] = '🇨🇲  CAMERAJOB WATCH — Rapport de Veille'
    ws['A1'].font = Font(name='Calibri', bold=True, size=18, color=BLANC)
    ws['A1'].fill = PatternFill("solid", fgColor=VERT)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}  |  {len(offres)} offres exportées"
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='555555')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].fill = PatternFill("solid", fgColor=VERT_L)
    ws.row_dimensions[2].height = 20

    # ── Statistiques ──
    emplois = [o for o in offres if o.type_offre == 'emploi']
    appels  = [o for o in offres if o.type_offre == 'appel_offre']

    # Statistiques par secteur
    secteurs = {}
    for o in offres:
        s = o.secteur or 'Non classifié'
        secteurs[s] = secteurs.get(s, 0) + 1

    # Statistiques par source
    sources = {}
    for o in offres:
        s = o.source or 'Inconnue'
        sources[s] = sources.get(s, 0) + 1

    # Statistiques par ville
    villes = {}
    for o in offres:
        v = o.localisation or 'Cameroun'
        villes[v] = villes.get(v, 0) + 1

    row = 4
    stats_data = [
        ('📊 STATISTIQUES GÉNÉRALES', '', '', ''),
        ('Total des offres exportées', len(offres), 'Sources différentes', len(sources)),
        ("Offres d'emploi", len(emplois), "Appels d'offres", len(appels)),
        ('Secteurs représentés', len(secteurs), 'Villes/Régions', len(villes)),
    ]

    for i, (a, b, c, d) in enumerate(stats_data):
        ws.row_dimensions[row + i].height = 25
        ws.cell(row=row+i, column=1, value=a)
        ws.cell(row=row+i, column=2, value=b)
        ws.cell(row=row+i, column=3, value=c)
        ws.cell(row=row+i, column=4, value=d)

        if i == 0:
            for col in [1, 2, 3, 4]:
                ws.cell(row=row+i, column=col).fill = PatternFill("solid", fgColor=VERT_C)
                ws.cell(row=row+i, column=col).font = Font(bold=True, color=BLANC, size=12)
        else:
            for col in [1, 3]:
                ws.cell(row=row+i, column=col).font = Font(bold=True, size=10)
                ws.cell(row=row+i, column=col).fill = PatternFill("solid", fgColor=GRIS)
            for col in [2, 4]:
                ws.cell(row=row+i, column=col).font = Font(bold=True, size=14, color=VERT)
                ws.cell(row=row+i, column=col).alignment = Alignment(horizontal='center')

    # ── Top secteurs ──
    row = 10
    ws.cell(row=row, column=1, value='🏷️ TOP SECTEURS').font = Font(bold=True, size=12, color=BLANC)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=ROUGE)
    ws.merge_cells(f'A{row}:B{row}')

    ws.cell(row=row, column=3, value='🏢 TOP SOURCES').font = Font(bold=True, size=12, color=BLANC)
    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=BLEU)
    ws.merge_cells(f'C{row}:D{row}')

    row += 1
    top_secteurs = sorted(secteurs.items(), key=lambda x: x[1], reverse=True)[:10]
    top_sources  = sorted(sources.items(),  key=lambda x: x[1], reverse=True)[:10]

    for i, ((sec, nb), (src, nb2)) in enumerate(zip(
        top_secteurs + [('', '')] * 10,
        top_sources  + [('', '')] * 10
    )):
        if i >= 10: break
        bg = GRIS_F if i % 2 == 0 else BLANC
        fill = PatternFill("solid", fgColor=bg)

        c1 = ws.cell(row=row+i, column=1, value=sec)
        c2 = ws.cell(row=row+i, column=2, value=nb if nb else '')
        c3 = ws.cell(row=row+i, column=3, value=src)
        c4 = ws.cell(row=row+i, column=4, value=nb2 if nb2 else '')

        for c in [c1, c2, c3, c4]:
            c.fill = fill
            c.font = Font(size=10)
        c2.alignment = Alignment(horizontal='center')
        c4.alignment = Alignment(horizontal='center')
        c2.font = Font(bold=True, size=11, color=VERT)
        c4.font = Font(bold=True, size=11, color=BLEU)


def _creer_onglet_emplois(ws, emplois):
    """Onglet dédié aux offres d'emploi"""
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 50

    # Titre
    ws.merge_cells('A1:I1')
    ws['A1'] = f"💼 OFFRES D'EMPLOI — {len(emplois)} offres  |  CamerJob Watch"
    ws['A1'].font = Font(bold=True, size=14, color=BLANC)
    ws['A1'].fill = PatternFill("solid", fgColor=VERT)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # En-têtes
    headers = ['N°', 'Titre du Poste', 'Organisation', 'Localisation',
               'Secteur', 'Type Contrat', 'Date Limite', 'Source', 'Lien / URL']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header(ws, 2, len(headers), bg=VERT_C)
    ws.row_dimensions[2].height = 30

    # Données
    for i, offre in enumerate(emplois):
        row = i + 3
        style_row_data(ws, row, len(headers), alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 40

        ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=2, value=offre.titre or '').alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=3, value=offre.organisation or '').alignment = Alignment(vertical='top')
        ws.cell(row=row, column=4, value=offre.localisation or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=5, value=offre.secteur or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=6, value=offre.type_contrat or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=7, value=offre.date_limite or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=8, value=offre.source or '').alignment = Alignment(vertical='top')

        # URL cliquable
        if offre.url:
            cell_url = ws.cell(row=row, column=9)
            cell_url.value = 'Voir l\'offre →'
            cell_url.hyperlink = offre.url
            cell_url.font = Font(color='1565C0', underline='single', size=10)
        else:
            ws.cell(row=row, column=9, value='N/A')

    # Filtre automatique
    ws.auto_filter.ref = f'A2:I{len(emplois)+2}'
    ws.freeze_panes = 'A3'


def _creer_onglet_appels_offres(ws, appels):
    """Onglet dédié aux appels d'offres"""
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 50

    # Titre
    ws.merge_cells('A1:H1')
    ws['A1'] = f"📋 APPELS D'OFFRES — {len(appels)} marchés  |  CamerJob Watch"
    ws['A1'].font = Font(bold=True, size=14, color=BLANC)
    ws['A1'].fill = PatternFill("solid", fgColor=ROUGE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # En-têtes
    headers = ['N°', "Intitulé du Marché", 'Autorité / Organisation',
               'Localisation', 'Secteur', 'Date Publication',
               'Date Limite', 'Lien / URL']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header(ws, 2, len(headers), bg=ROUGE)
    ws.row_dimensions[2].height = 30

    # Données
    for i, offre in enumerate(appels):
        row = i + 3
        style_row_data(ws, row, len(headers), alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 45

        ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=2, value=offre.titre or '').alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=3, value=offre.organisation or '').alignment = Alignment(vertical='top')
        ws.cell(row=row, column=4, value=offre.localisation or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=5, value=offre.secteur or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=6, value=offre.date_publication or '').alignment = Alignment(horizontal='center', vertical='top')

        # Date limite en rouge si proche
        cell_date = ws.cell(row=row, column=7, value=offre.date_limite or '')
        cell_date.alignment = Alignment(horizontal='center', vertical='top')
        cell_date.font = Font(color=ROUGE, bold=True, size=10) if offre.date_limite else Font(size=10)

        # URL cliquable
        if offre.url:
            cell_url = ws.cell(row=row, column=8)
            cell_url.value = 'Voir le marché →'
            cell_url.hyperlink = offre.url
            cell_url.font = Font(color='1565C0', underline='single', size=10)
        else:
            ws.cell(row=row, column=8, value='N/A')

    ws.auto_filter.ref = f'A2:H{len(appels)+2}'
    ws.freeze_panes = 'A3'


def _creer_onglet_toutes(ws, offres):
    """Onglet avec toutes les offres"""
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 40

    ws.merge_cells('A1:J1')
    ws['A1'] = f"📋 TOUTES LES OFFRES — {len(offres)} résultats  |  CamerJob Watch"
    ws['A1'].font = Font(bold=True, size=14, color=BLANC)
    ws['A1'].fill = PatternFill("solid", fgColor=NOIR)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    headers = ['N°', 'Type', 'Titre', 'Organisation', 'Localisation',
               'Secteur', 'Date Limite', 'Source', 'Collecté le', 'URL']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header(ws, 2, len(headers), bg='333333')
    ws.row_dimensions[2].height = 28

    for i, offre in enumerate(offres):
        row = i + 3
        style_row_data(ws, row, len(headers), alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 35

        type_cell = ws.cell(row=row, column=2)
        if offre.type_offre == 'emploi':
            type_cell.value = '💼 Emploi'
            type_cell.font = Font(color=VERT, bold=True, size=9)
        else:
            type_cell.value = '📋 AO'
            type_cell.font = Font(color=ROUGE, bold=True, size=9)
        type_cell.alignment = Alignment(horizontal='center', vertical='top')

        ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=3, value=offre.titre or '').alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=4, value=offre.organisation or '').alignment = Alignment(vertical='top')
        ws.cell(row=row, column=5, value=offre.localisation or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=6, value=offre.secteur or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=7, value=offre.date_limite or '').alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row, column=8, value=offre.source or '').alignment = Alignment(vertical='top')
        ws.cell(row=row, column=9, value=offre.cree_le.strftime('%d/%m/%Y') if offre.cree_le else '').alignment = Alignment(horizontal='center', vertical='top')

        if offre.url:
            cell_url = ws.cell(row=row, column=10)
            cell_url.value = offre.url[:80] + '...' if len(offre.url or '') > 80 else offre.url
            cell_url.hyperlink = offre.url
            cell_url.font = Font(color='1565C0', underline='single', size=9)

    ws.auto_filter.ref = f'A2:J{len(offres)+2}'
    ws.freeze_panes = 'A3'
